"""Parser for MCA Snapshot (Daily Snapshot) report family.

Source layout (single sheet "Summary"):
  Row 2 = title context (client, BRM info)
  Row 3 = headers: Schedule, Volume, Client, Region, District, CID, Store Name,
    RVP, Team, Total Active Requests, Total Remaining Requests,
    Mon-Sun Build Total, WTD Build, Buildup Percentage, Last Build Date,
    Next Event, Last Crew Build, Last Manager Walk Date, Days Since Last Manager Walk,
    MTD Walks, MTD No Work Events, Open Invoice Amount, Missed Events, Late SCEs,
    Monthly Completed SCEs, On Time %, Approved Incentive Events,
    Last FS Quality Walk Date, Days Since Last FS Quality Walk, Permanent Schedule,
    Total Weekly SCEs, MTD Damage Reports, WTD Invoice Amount, LY WTD Invoice Amount,
    MTD Invoice Amount, LY MTD Invoice Amount, YTD Invoice Amount, LY YTD Invoice Amount,
    WTD Asset Invoice Amount, MTD Asset Invoice Amount, YTD Asset Invoice Amount,
    MTD Sub Technician Invoice Amount, MTD Commission Invoice Estimate,
    LOW Last Week Grill Sales, WTD Grill Builds

Each row = one CID/store's daily snapshot on a given date.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl

from .base import BaseParser
from ..warehouse.repository import WarehouseRepository

logger = logging.getLogger(__name__)

SNAPSHOT_FILENAME_PATTERN = re.compile(
    r"Daily\s+Snapshot\s+(\d{1,2})\.(\d{1,2})\.(\d{2,4})",
    re.IGNORECASE,
)

EXPECTED_HEADERS = {"cid", "store name", "team", "total active requests", "wtd build"}

COLUMN_MAP = {
    "Schedule": "schedule",
    "Volume": "volume",
    "Client": "client",
    "Region": "region",
    "District": "district",
    "CID": "cid",
    "Store Name": "store_name",
    "RVP": "rvp",
    "Team": "team",
    "Total Active Requests": "total_active_requests",
    "Total Remaining Requests": "total_remaining_requests",
    "Mon Build Total": "mon_build",
    "Tues Build Total": "tues_build",
    "Wed Build Total": "wed_build",
    "Thurs Build Total": "thurs_build",
    "Fri Build Total": "fri_build",
    "Sat Build Total": "sat_build",
    "Sun Build Total": "sun_build",
    "WTD Build": "wtd_build",
    "Buildup Percentage": "buildup_pct",
    "Last Build Date": "last_build_date",
    "Next Event": "next_event",
    "Last Crew Build": "last_crew_build",
    "Last Manager Walk Date": "last_manager_walk",
    "Days Since Last Manager Walk": "days_since_manager_walk",
    "MTD Walks": "mtd_walks",
    "MTD No Work Events": "mtd_no_work_events",
    "Open Invoice Amount": "open_invoice_amount",
    "Missed Events": "missed_events",
    "Late SCEs": "late_sces",
    "Monthly Completed SCEs": "monthly_completed_sces",
    "On Time %": "on_time_pct",
    "Approved Incentive Events": "approved_incentive_events",
    "Permanent Schedule": "permanent_schedule",
    "Total Weekly SCEs": "total_weekly_sces",
    "MTD Damage Reports": "mtd_damage_reports",
    "WTD Invoice Amount": "wtd_invoice_amount",
    "LY WTD Invoice Amount": "ly_wtd_invoice_amount",
    "MTD Invoice Amount": "mtd_invoice_amount",
    "LY MTD Invoice Amount": "ly_mtd_invoice_amount",
    "YTD Invoice Amount": "ytd_invoice_amount",
    "LY YTD Invoice Amount": "ly_ytd_invoice_amount",
    "WTD Asset Invoice Amount": "wtd_asset_invoice_amount",
    "MTD Asset Invoice Amount": "mtd_asset_invoice_amount",
    "YTD Asset Invoice Amount": "ytd_asset_invoice_amount",
    "MTD Sub Technician Invoice Amount": "mtd_sub_tech_invoice_amount",
    "MTD Commission Invoice Estimate": "mtd_commission_invoice_est",
}

NUMERIC_METRICS = [
    "total_active_requests", "total_remaining_requests",
    "wtd_build", "buildup_pct", "days_since_manager_walk",
    "mtd_walks", "mtd_no_work_events", "open_invoice_amount",
    "missed_events", "late_sces", "monthly_completed_sces",
    "on_time_pct", "total_weekly_sces", "mtd_damage_reports",
    "wtd_invoice_amount", "ly_wtd_invoice_amount",
    "mtd_invoice_amount", "ly_mtd_invoice_amount",
    "ytd_invoice_amount", "ly_ytd_invoice_amount",
]


class MCASnapshotParser(BaseParser):

    @property
    def family_name(self) -> str:
        return "mca_snapshot"

    def can_parse(self, filepath: str) -> float:
        path = Path(filepath)
        name = path.name.lower()

        if path.suffix.lower() not in (".xlsx", ".xlsm", ".xls", ".xlsb"):
            return 0.0

        if SNAPSHOT_FILENAME_PATTERN.search(path.name):
            return 0.9

        if "daily snapshot" in name or "mca snapshot" in name:
            return 0.7

        if "snapshot" in name:
            return 0.4

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                if sheet.lower() in ("summary",):
                    ws = wb[sheet]
                    for row in ws.iter_rows(min_row=1, max_row=5, max_col=51, values_only=True):
                        headers = {str(v).lower().strip() for v in row if v}
                        if len(headers & EXPECTED_HEADERS) >= 3:
                            wb.close()
                            return 0.6
            wb.close()
        except Exception:
            pass
        return 0.0

    def _extract_date_from_filename(self, filepath: str) -> date | None:
        match = SNAPSHOT_FILENAME_PATTERN.search(Path(filepath).name)
        if match:
            m, d, y = int(match.group(1)), int(match.group(2)), int(match.group(3))
            if y < 100:
                y += 2000
            try:
                return date(y, m, d)
            except ValueError:
                pass
        return None

    def parse(self, filepath: str) -> List[Dict[str, Any]]:
        records = []
        snapshot_date = self._extract_date_from_filename(filepath)

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            logger.error("Cannot open %s: %s", filepath, e)
            return records

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_map = {}
            header_row = None

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=False), 1):
                cells = {str(c.value).strip(): c.column - 1 for c in row if c.value is not None}
                matched = sum(1 for k in cells if k in COLUMN_MAP)
                if matched >= 5:
                    for orig, col_idx in cells.items():
                        if orig in COLUMN_MAP:
                            header_map[COLUMN_MAP[orig]] = col_idx
                    header_row = row_idx
                    break

            if not header_map:
                continue

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(row):
                    continue
                record = {"_source_file": filepath, "_sheet": sheet_name,
                         "snapshot_date": snapshot_date}
                for field, col_idx in header_map.items():
                    if col_idx < len(row):
                        record[field] = row[col_idx]
                if record.get("cid") is not None:
                    records.append(record)

        wb.close()
        logger.info("Parsed %d snapshot records from %s", len(records), Path(filepath).name)
        return records

    def normalize_and_store(self, records: List[Dict], repo: WarehouseRepository,
                           source_file_id: int, run_id: int) -> int:
        family = repo.get_or_create_family(
            "mca_snapshot",
            "MCA Daily Snapshot - CID/store-level operational snapshot"
        )

        metric_cache = {}
        for m in NUMERIC_METRICS:
            metric_cache[m] = repo.get_or_create_metric(m, m.replace("_", " ").title())

        obs_count = 0
        for rec in records:
            snap_date = rec.get("snapshot_date")
            if not isinstance(snap_date, date):
                continue

            cid = rec.get("cid")
            if cid is None:
                continue

            cid_entity = repo.get_or_create_entity("cid", str(cid))
            store_name = rec.get("store_name")
            if store_name:
                repo.add_entity_alias(cid_entity.id, str(store_name), "mca_snapshot")

            team_id = None
            team_name = rec.get("team")
            if team_name:
                team_entity = repo.get_or_create_entity("team", str(team_name).strip())
                team_id = team_entity.id
                if cid_entity.parent_id != team_id:
                    cid_entity.parent_id = team_id

            rvp_name = rec.get("rvp")
            if rvp_name:
                repo.get_or_create_entity("rvp", str(rvp_name).strip())

            region = rec.get("region")
            if region:
                repo.get_or_create_entity("region", str(region).strip())

            period = repo.get_or_create_period(
                "day", snap_date, snap_date,
                label=snap_date.strftime("%Y-%m-%d"),
            )

            for field in NUMERIC_METRICS:
                val = rec.get(field)
                if val is not None:
                    try:
                        fval = float(val)
                    except (ValueError, TypeError):
                        continue
                    metric = metric_cache[field]
                    repo.add_observation(
                        entity_id=cid_entity.id, period_id=period.id,
                        metric_id=metric.id, value=fval,
                        parent_entity_id=team_id,
                        report_family_id=family.id,
                        source_file_id=source_file_id,
                        source_sheet=rec.get("_sheet"),
                        ingest_run_id=run_id,
                    )
                    obs_count += 1

        repo.session.flush()
        return obs_count
