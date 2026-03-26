"""Parser for Daily Technician Performance report family.

Source layout (single sheet "Report"):
  Row 1 = headers: Client, CID, Location Name, Location RVP, Location Team,
    Technician Team, Technician, Role, Commission Rate, Crew Work, Date, Hours,
    First Checkin, Last Checkout, Estimated Hours, Difference, Units, Amount,
    Units/Hour, Dollar/Hour, Allow Stickers, Force Stickers Audit, Daily Units,
    WTD Units, Estimated Sliding Scale Rate

Each row = one technician's work at one store on one date.
"""

from __future__ import annotations

import logging
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
import pandas as pd

from .base import BaseParser
from ..warehouse.repository import WarehouseRepository

logger = logging.getLogger(__name__)

DTP_FILENAME_PATTERN = re.compile(
    r"Daily\s+Technician\s+Performance\s+(\d{1,2})\.(\d{1,2})\.(\d{2,4})",
    re.IGNORECASE,
)

EXPECTED_COLUMNS = {
    "client", "cid", "technician team", "technician", "role",
    "commission rate", "date", "hours", "units", "amount",
}

COLUMN_MAP = {
    "Client": "client",
    "CID": "cid",
    "Location Name": "location_name",
    "Location RVP": "location_rvp",
    "Location Team": "location_team",
    "Technician Team": "technician_team",
    "Technician": "technician",
    "Role": "role",
    "Commission Rate": "commission_rate",
    "Crew Work": "crew_work",
    "Date": "work_date",
    "Hours": "hours",
    "First Checkin": "first_checkin",
    "Last Checkout": "last_checkout",
    "Estimated Hours": "estimated_hours",
    "Difference": "difference",
    "Units": "units",
    "Amount": "amount",
    "Units/Hour": "units_per_hour",
    "Dollar/Hour": "dollars_per_hour",
    "Allow Stickers": "allow_stickers",
    "Force Stickers Audit": "force_stickers_audit",
    "Daily Units": "daily_units",
    "WTD Units": "wtd_units",
    "Estimated Sliding Scale Rate": "est_sliding_scale_rate",
}


class DailyTechPerformanceParser(BaseParser):

    @property
    def family_name(self) -> str:
        return "daily_tech_performance"

    def can_parse(self, filepath: str) -> float:
        path = Path(filepath)
        name = path.name.lower()

        if path.suffix.lower() not in (".xlsx", ".xlsm", ".xls", ".xlsb"):
            return 0.0

        if DTP_FILENAME_PATTERN.search(path.name):
            return 0.9

        if "technician performance" in name or "tech performance" in name:
            return 0.7

        if "daily" in name and "performance" in name:
            return 0.5

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                for row in ws.iter_rows(min_row=1, max_row=3, max_col=25, values_only=True):
                    headers = {str(v).lower().strip() for v in row if v}
                    if len(headers & EXPECTED_COLUMNS) >= 5:
                        wb.close()
                        return 0.6
            wb.close()
        except Exception:
            pass
        return 0.0

    def parse(self, filepath: str) -> List[Dict[str, Any]]:
        path = Path(filepath)
        records = []

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            logger.error("Cannot open %s: %s", filepath, e)
            return records

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_map = {}
            header_row = None

            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=False), 1):
                cells = {str(c.value).strip(): c.column - 1 for c in row if c.value is not None}
                matched = sum(1 for k in cells if k in COLUMN_MAP)
                if matched >= 5:
                    header_map = {}
                    for orig, col_idx in cells.items():
                        if orig in COLUMN_MAP:
                            header_map[COLUMN_MAP[orig]] = col_idx
                    header_row = row_idx
                    break

            if not header_map:
                continue

            for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
                if not any(row):
                    continue
                record = {"_source_file": filepath, "_sheet": sheet_name}
                for field, col_idx in header_map.items():
                    if col_idx < len(row):
                        record[field] = row[col_idx]
                if record.get("technician") and record.get("hours") is not None:
                    self._clean_record(record, path)
                    records.append(record)

        wb.close()
        logger.info("Parsed %d records from %s", len(records), path.name)
        return records

    def _clean_record(self, record: Dict, source_path: Path) -> None:
        if isinstance(record.get("work_date"), datetime):
            record["work_date"] = record["work_date"].date()
        elif isinstance(record.get("work_date"), str):
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                try:
                    record["work_date"] = datetime.strptime(record["work_date"], fmt).date()
                    break
                except ValueError:
                    continue

        if record.get("work_date") is None:
            match = DTP_FILENAME_PATTERN.search(source_path.name)
            if match:
                m, d, y = int(match.group(1)), int(match.group(2)), int(match.group(3))
                if y < 100:
                    y += 2000
                try:
                    record["work_date"] = date(y, m, d)
                except ValueError:
                    pass

        for field in ("hours", "units", "amount", "commission_rate",
                      "units_per_hour", "dollars_per_hour"):
            val = record.get(field)
            if val is not None:
                try:
                    record[field] = float(val)
                except (ValueError, TypeError):
                    record[field] = 0.0

    def normalize_and_store(self, records: List[Dict], repo: WarehouseRepository,
                           source_file_id: int, run_id: int) -> int:
        family = repo.get_or_create_family(
            "daily_tech_performance",
            "Daily Technician Performance - technician-level daily work records"
        )

        metrics = {
            "hours": repo.get_or_create_metric("hours", "Hours Worked", "hours", "sum"),
            "units": repo.get_or_create_metric("units", "Units Completed", "count", "sum"),
            "amount": repo.get_or_create_metric("gross_revenue", "Gross Revenue", "dollars", "sum"),
            "commission_rate": repo.get_or_create_metric("commission_rate", "Commission Rate", "rate", "avg"),
            "units_per_hour": repo.get_or_create_metric("units_per_hour", "Units Per Hour", "rate", "avg"),
            "dollars_per_hour": repo.get_or_create_metric("gross_dollars_per_hour", "Gross $/Hr", "dollars/hour", "avg"),
        }

        obs_count = 0
        for rec in records:
            work_date = rec.get("work_date")
            if not isinstance(work_date, date):
                continue

            tech_name = str(rec.get("technician", "")).strip()
            team_name = str(rec.get("technician_team", "")).strip()
            if not tech_name:
                continue

            tech_entity = repo.get_or_create_entity("technician", tech_name)
            period = repo.get_or_create_period(
                "day", work_date, work_date,
                label=work_date.strftime("%Y-%m-%d"),
            )

            team_id = None
            if team_name:
                team_entity = repo.get_or_create_entity("team", team_name)
                team_id = team_entity.id
                if tech_entity.parent_id != team_id:
                    tech_entity.parent_id = team_id

            for field, metric in metrics.items():
                val = rec.get(field)
                if val is not None:
                    try:
                        fval = float(val)
                    except (ValueError, TypeError):
                        continue
                    repo.add_observation(
                        entity_id=tech_entity.id, period_id=period.id,
                        metric_id=metric.id, value=fval,
                        parent_entity_id=team_id,
                        report_family_id=family.id,
                        source_file_id=source_file_id,
                        source_sheet=rec.get("_sheet"),
                        ingest_run_id=run_id,
                    )
                    obs_count += 1

            cid = rec.get("cid")
            loc_name = rec.get("location_name")
            if cid is not None:
                cid_entity = repo.get_or_create_entity("cid", str(cid))
                if loc_name:
                    repo.add_entity_alias(cid_entity.id, str(loc_name), "daily_tech_performance")

            rvp = rec.get("location_rvp")
            if rvp:
                repo.get_or_create_entity("rvp", str(rvp).strip())

        repo.session.flush()
        return obs_count
